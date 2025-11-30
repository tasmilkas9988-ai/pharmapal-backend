#!/usr/bin/env python3
"""
Import medications from Excel file to MongoDB
Supports bilingual data (English & Arabic)
"""
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv
from pathlib import Path
import openpyxl
import uuid

async def import_medications():
    # Load environment
    ROOT_DIR = Path("/app/backend")
    load_dotenv(ROOT_DIR / '.env')
    
    mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
    client = AsyncIOMotorClient(mongo_url)
    db = client[os.environ.get('DB_NAME', 'pharmapal_db')]
    
    # Load Excel file
    print("📂 جاري تحميل الملف...")
    wb = openpyxl.load_workbook('/tmp/medications.xlsx', data_only=True)
    ws = wb.active
    
    # Clear existing medications
    print("🗑️  جاري حذف قاعدة البيانات القديمة...")
    result = await db.sfda_medications.delete_many({})
    print(f"✅ تم حذف {result.deleted_count} دواء قديم")
    
    # Import new medications
    medications = []
    errors = 0
    
    print("📥 جاري استيراد الأدوية...")
    for row_idx in range(2, ws.max_row + 1):
        try:
            row = list(ws[row_idx])
            
            # Extract data
            trade_name_en = str(row[0].value).strip() if row[0].value else ""
            trade_name_ar = str(row[1].value).strip() if row[1].value else trade_name_en
            scientific_name_en = str(row[2].value).strip() if row[2].value else ""
            scientific_name_ar = str(row[3].value).strip() if row[3].value else scientific_name_en
            package_size = row[4].value if row[4].value else 1
            strength = str(row[6].value) if row[6].value else ""
            price = float(row[8].value) if row[8].value else 0.0
            package_type_en = str(row[10].value).strip() if row[10].value else ""
            package_type_ar = str(row[11].value).strip() if row[11].value else package_type_en
            strength_unit = str(row[12].value).strip() if row[12].value else ""
            admin_route_en = str(row[14].value).strip() if row[14].value else ""
            admin_route_ar = str(row[15].value).strip() if row[15].value else admin_route_en
            dosage_form_en = str(row[16].value).strip() if row[16].value else ""
            dosage_form_ar = str(row[17].value).strip() if row[17].value else dosage_form_en
            storage_en = str(row[18].value).strip() if row[18].value else ""
            storage_ar = str(row[19].value).strip() if row[19].value else storage_en
            manufacturer_en = str(row[20].value).strip() if row[20].value else ""
            manufacturer_ar = str(row[21].value).strip() if row[21].value else manufacturer_en
            legal_status_en = str(row[22].value).strip() if row[22].value else "Prescription"
            legal_status_ar = str(row[23].value).strip() if row[23].value else "يحتاج الى وصفة طبية"
            
            # Skip if no trade name
            if not trade_name_en or trade_name_en.lower() == 'nan':
                continue
            
            # Create medication document
            medication = {
                "id": str(uuid.uuid4()),
                # Trade name (bilingual)
                "trade_name": trade_name_en,
                "trade_name_ar": trade_name_ar,
                "trade_name_lower": trade_name_en.lower(),
                
                # Scientific name (bilingual)
                "scientific_name": scientific_name_en,
                "scientific_name_ar": scientific_name_ar,
                "active_ingredients": scientific_name_en,
                "active_ingredients_lower": scientific_name_en.lower(),
                
                # Package info
                "package_size": int(package_size) if isinstance(package_size, (int, float)) else 1,
                "pack": package_type_en,
                "pack_ar": package_type_ar,
                
                # Strength
                "strength": strength,
                "strength_unit": strength_unit,
                
                # Price
                "price_sar": price,
                
                # Dosage form (bilingual)
                "dosage_form": dosage_form_en,
                "dosage_form_ar": dosage_form_ar,
                
                # Administration route (bilingual)
                "administration_route": admin_route_en,
                "administration_route_ar": admin_route_ar,
                
                # Storage (bilingual)
                "storage_conditions": storage_en,
                "storage_conditions_ar": storage_ar,
                
                # Manufacturer (bilingual)
                "manufacturer": manufacturer_en,
                "manufacturer_ar": manufacturer_ar,
                
                # Legal status (bilingual)
                "legal_status": legal_status_en,
                "legal_status_en": legal_status_en,
                "legal_status_ar": legal_status_ar,
                
                # Source
                "source": "SFDA_2025_BILINGUAL"
            }
            
            medications.append(medication)
            
            # Batch insert every 1000 records
            if len(medications) >= 1000:
                await db.sfda_medications.insert_many(medications)
                print(f"✅ تم استيراد {row_idx - 1} دواء...")
                medications = []
                
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"⚠️ خطأ في الصف {row_idx}: {str(e)}")
    
    # Insert remaining medications
    if medications:
        await db.sfda_medications.insert_many(medications)
    
    # Create indexes
    print("\n🔍 جاري إنشاء الفهارس...")
    await db.sfda_medications.create_index("trade_name_lower")
    await db.sfda_medications.create_index("active_ingredients_lower")
    await db.sfda_medications.create_index([("trade_name_lower", "text"), ("active_ingredients_lower", "text")])
    
    # Get final count
    total = await db.sfda_medications.count_documents({})
    
    print(f"\n✅ اكتمل الاستيراد!")
    print(f"📊 إجمالي الأدوية: {total}")
    print(f"⚠️ أخطاء: {errors}")
    
    client.close()

if __name__ == "__main__":
    asyncio.run(import_medications())
